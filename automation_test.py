import datetime
import openpyxl as pyxl
from openpyxl.styles import Font


wb = pyxl.Workbook()
ws = wb.active

ws.title = "First Page"
ws.sheet_properties.tabColor = "1072BA"

ws1 = wb.create_sheet("Mysheet")



ws["A1"] = 1
ws["D8"] = 2
ws["G7"] = 3
ws["D1"] = "=A1+D8+G7"



ws["B2"] = datetime.datetime(2022, 8, 28, 19, 54, 44)

wb.save("Automatetest.xlsx")

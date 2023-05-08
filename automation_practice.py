import datetime
import openpyxl as xl
from pathlib import Path


def none():

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "First Page"
    ws.sheet_properties.tabColor = "1072BA"
    ws1 = wb.create_sheet("Mysheet")
    ws["A1"] = 1
    ws["D8"] = 2
    ws["G7"] = 3
    ws["D1"] = "=A1+D8+G7"
    ws["B2"] = datetime.datetime.now()
    wb.save("Automatetest.xlsx")


def half_the_value(filename):

    wbook = xl.load_workbook(filename)
    for sheet in wbook.worksheets:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            price_halved = cell.value / 2
            price_halved_cell = sheet.cell(row, 4)
            price_halved_cell.value = price_halved

    wbook.save(filename)

def find_files():

    path = Path()
    for file in path.glob("*.xlsx"):
        print(file)

find_files()

my_file = Path("nt1.xlsx")
if my_file.is_file():
    half_the_value(my_file)
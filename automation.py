from pathlib import Path
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    sheet["d1"] = "-10%"

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
                min_row=2, 
                max_row=sheet.max_row,
                min_col=4,
                max_col=4
                )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "f2")

    wb.save(filename)

def find_files():

    path = Path()
    for file in path.glob("*"):
        print(file)

#find_files()

my_file = Path("transactionscopy.xlsx")
if my_file.is_file():
    process_workbook(my_file)







